import json
from openpyxl import load_workbook
from openpyxl.styles import *
from PyQt5.QtCore import QThread, pyqtSignal
from .. import PostgisDB
from .config import Configurer


class CuttingAreaExport(QThread):
    """
    Поток для чтения данных из базы и записи в JSON.
    """

    signal_message_result = pyqtSignal(str)
    CUTTING_AREAS_TABLES = {
        "area": "uid",  # Главная таблица для экспорта в xlsx
        "area_data": "area_uid",
        "area_line": "uid",
        "area_points": "area_uid",
        "trees": "area_uuid",
        "trees_not_cutting": "area_uuid",
        "trees_trf_height": "area_uuid",
    }

    def __init__(
        self, uuid_list: list, path_to_file: str, file_extension: str
    ):
        QThread.__init__(self)
        self.uuid_list = uuid_list
        self.path_to_file = path_to_file

        if file_extension == "json":
            self.run = self.write_to_json
        elif file_extension == "xlsx":
            self.run = self.write_to_xlsx

    def get_data_from_db(self):
        """
        Получение данных из БД для прописанных таблиц и полей идентификатора

        Returns:
            [type]: [Данные для сохранения в JSON]
        """
        output_data = {}

        try:
            for table_name in self.CUTTING_AREAS_TABLES:
                field_name = self.CUTTING_AREAS_TABLES.get(table_name)

                sql_data = PostgisDB.PostGisDB().getQueryResult(
                    f"select * from {table_name} where {field_name} in ({str(self.uuid_list)[1:-1]})",
                    as_dict=True,
                )
                if sql_data != [{}]:  # Если таблица не пуста
                    output_data.setdefault(table_name, sql_data)

        except Exception as mes:
            self.signal_message_result.emit(str(mes))

        return output_data

    def write_to_json(self) -> None:
        """
        Запись выбранных данных в JSON файл
        """
        with open(self.path_to_file, "w") as outfile:
            json.dump(
                self.get_data_from_db(), outfile, ensure_ascii=False, indent=2
            )
        self.signal_message_result.emit(
            f"Лесосеки успешно экспортированы в:\n{self.path_to_file}"
        )

    def write_to_xlsx(self) -> None:

        thin_border = borders.Border(
            left=borders.Side(style="thin"),
            right=borders.Side(style="thin"),
            top=borders.Side(style="thin"),
            bottom=borders.Side(style="thin"),
        )
        try:
            template_file = Configurer.resolve(None, "res/cutting_areas.xlsx")
            workbook = load_workbook(template_file)
            ws = workbook["Экспорт лесосек"]

            main_data = self.get_data_from_db().get("area", False)
            if main_data:
                for num, area in enumerate(main_data):
                    num += 2

                    ws.cell(
                        row=num, column=1, value=area.get("leshos_text", "—")
                    ).alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=2, value=area.get("lesnich_text", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=3, value=area.get("num_kv", "—")
                    ).alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=4, value=area.get("num_vds", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=5, value=area.get("num", "—")
                    ).alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=6, value=area.get("usetype", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=7, value=area.get("cuttingtyp", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=8, value=area.get("area", "—")
                    ).alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=9, value=area.get("date", "—")
                    ).alignment = Alignment(
                        wrap_text=False,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=10, value=area.get("fio", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    ws.cell(
                        row=num, column=11, value=area.get("info", "—")
                    ).alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="center",
                        horizontal="center",
                    )
                    cell.border = thin_border
            workbook.save(self.path_to_file)
            self.signal_message_result.emit(
                f"Лесосеки успешно экспортированы в:\n{self.path_to_file}"
            )
        except Exception as e:
            self.signal_message_result.emit(str(e))
