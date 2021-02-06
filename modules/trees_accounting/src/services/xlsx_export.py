from PyQt5.QtGui import QStandardItemModel
from PyQt5.QtCore import QThread, pyqtSignal
from ..models.dictionary import Species, KindSeeds
from ..services.config import BasicDir
from openpyxl import load_workbook


class XLSXExportedData(QThread):
    """
    ! Не обработаны ошибки
    Поток для сохранения данных в JSON для импорта в АРМ Лесопользование.
    """

    signal_message_result = pyqtSignal(dict)

    def __init__(
        self,
        att_data: dict,
        model_liquid: QStandardItemModel,
        model_not_cutting: QStandardItemModel,
        export_file: str,
    ):
        QThread.__init__(self)
        self.att_data = att_data
        self.model_liquid = model_liquid
        self.model_not_cutting = model_not_cutting
        self.export_file = export_file

    def write_trees_liquid_data(self, wb):
        """Добавляю данные в книгу Перечет"""
        species_name = {}
        ws = wb["Перечет"]
        instance_list = self.model_liquid.as_list()

        if type(instance_list) is list:
            for instance in instance_list:

                # Формула для получения номера столбца в xlsx для нужной породы,
                # а так же для (Деловых, Дровяных(Соответственно + 1)):
                # По 3 строке это будет название породы
                xlsx_column = (
                    self.model_liquid.species_position(instance["code_species"]) + 2
                )
                if xlsx_column > 12:
                    xlsx_column += 1  # это будет новая страница

                # Формула для получения номера строки в xlsx для нужного диаметра:
                xlsx_row = self.model_liquid.row_number_by_dmr(instance["dmr"]) + 1

                # Получаю название породы по коду породы, если ещё не получал:
                if instance["code_species"] not in species_name.keys():
                    spc_name = Species.get(
                        Species.code_species == instance["code_species"]
                    ).name_species
                    species_name[instance["code_species"]] = spc_name
                    ws.cell(row=3, column=xlsx_column, value=spc_name)

                # Заношу значение в xlsx таблицу:
                if instance["num_ind"] != "0":
                    ws.cell(
                        row=xlsx_row, column=xlsx_column, value=int(instance["num_ind"])
                    )

                if instance["num_fuel"] != "0":
                    ws.cell(
                        row=xlsx_row,
                        column=xlsx_column + 1,
                        value=int(instance["num_fuel"]),
                    )

            return True
        instance_list["detailed_text"] = " "
        self.signal_message_result.emit(instance_list)
        return False

    def write_trees_not_cutting_data(self, wb):
        ws = wb["Семенники"]
        instance_list = self.model_not_cutting.as_list()

        for counter in range(len(instance_list)):
            row = counter + 4
            current_species_name = (
                Species.select(Species.name_species)
                .where(Species.code_species == instance_list[counter]["code_species"])
                .get()
                .name_species
            )
            current_seed_type_name = (
                KindSeeds.select(KindSeeds.name_kind_seeds)
                .where(
                    KindSeeds.code_kind_seeds
                    == instance_list[counter]["seed_type_code"]
                )
                .get()
                .name_kind_seeds
            )

            ws.cell(row=row, column=1, value=current_species_name)
            ws.cell(row=row, column=4, value=current_seed_type_name)
            ws.cell(row=row, column=8, value=instance_list[counter]["seed_dmr"])
            ws.cell(row=row, column=10, value=instance_list[counter]["seed_count"])
            ws.cell(row=row, column=12, value=instance_list[counter]["seed_number"])

        return True

    def run(self):
        file = BasicDir.get_module_dir("templates/template_accounting.xlsx")
        workbook = load_workbook(file)

        if self.write_trees_liquid_data(
            wb=workbook
        ) and self.write_trees_not_cutting_data(wb=workbook):

            workbook.save(self.export_file)
            self.signal_message_result.emit(
                {
                    "main_text": "Данные успешно экспортированы в\n" + self.export_file,
                    "detailed_text": None,
                }
            )
