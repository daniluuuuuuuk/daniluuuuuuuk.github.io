from PyQt5 import QtCore

from ..config import BasicDir

from openpyxl import load_workbook


class Data(QtCore.QThread):
    signal_status = QtCore.pyqtSignal(dict)

    def __init__(self, **args):
        QtCore.QThread.__init__(self)
        # self.uuid = args['uuid']
        self.table = args["table"]
        self.att_data = args["att_data"]
        self.export_file = args["export_file"]
        self.data_mapping = {
            "species": [
                "B",
                "F",
                "J",
                "N",
                "R",
                "W",
                "AA",
                "AE",
                "AI",
                "AM",
            ],  # Строка 3
            "dmrs": {
                "8": "5",
                "12": "6",
                "16": "7",
                "20": "8",
                "24": "9",
                "28": "10",  # Столбец A
                "32": "11",
                "36": "12",
                "40": "13",
                "44": "14",
                "48": "15",
                "52": "16",
                "56": "17",
                "60": "18",
                "64": "19",
                "68": "20",
                "72": "21",
                "76": "22",
                "80": "23",
                "84": "24",
                "88": "25",
                "92": "26",
                "96": "27",
                "100": "28",
                "104": "29",
                "108": "30",
                "112": "31",
                "116": "32",
                "120": "33",
                "124": "34",
            },
            "alp": [
                "B",
                "C",
                "E",
                "F",
                "G",
                "I",
                "J",
                "K",
                "M",
                "N",
                "O",
                "Q",
                "R",
                "S",
                "U",
                "W",
                "X",
                "Z",
                "AA",
                "AB",
                "AD",
                "AE",
                "AF",
                "AH",
                "AI",
                "AJ",
                "AL",
                "AM",
                "AN",
                "AP",
            ],
        }

    def collect_trees_data(self):
        """Собираю перечётку из GUI таблицы"""
        origin_data = {}
        columns = self.table.columnCount()
        rows = self.table.rowCount()
        for column in range(columns):
            if self.table.item(0, column):  # Нахожу породу
                species = self.table.item(0, column).text()
                origin_data[species] = {}
                for row in range(rows):
                    num_ind = self.table.item(row, column + 1)
                    num_fuel = self.table.item(row, column + 2)
                    num_deadwood = self.table.item(row, column + 3)
                    if 1 < row < rows - 2:
                        dmr = self.table.item(row, 0).text()
                        if num_ind:
                            num_ind = int(num_ind.text())
                        else:
                            num_ind = ""
                        if num_fuel:
                            num_fuel = int(num_fuel.text())
                        else:
                            num_fuel = ""
                        if num_deadwood:
                            num_deadwood = int(num_deadwood.text())
                        else:
                            num_deadwood = ""

                        if num_fuel == "" and num_ind == "" and num_deadwood == "":
                            None
                        else:
                            origin_data[species][dmr] = {
                                "num_ind": num_ind,
                                "num_fuel": num_fuel,
                                "num_deadwood": num_deadwood,
                            }
        return origin_data

    def write_trees_data(self, wb):
        """Записываю книгу Перечёт"""
        origin_data = self.collect_trees_data()
        ws = wb["Перечет"]
        species = list(origin_data.keys())

        for i in range(len(species)):
            ws[self.data_mapping["species"][i] + "3"] = species[i]
            sp_data = origin_data[species[i]]
            sp_dmrs = list(sp_data.keys())
            for u in range(len(sp_dmrs)):
                current_dmr = sp_dmrs[u]
                current_dmr_data = sp_data[sp_dmrs[u]]

                current_dmr_row = self.data_mapping["dmrs"][current_dmr]
                current_dmr_coll_ind = self.data_mapping["alp"][0]
                current_dmr_coll_fuel = self.data_mapping["alp"][1]
                current_dmr_coll_deadwood = self.data_mapping["alp"][2]

                ws[current_dmr_coll_ind + current_dmr_row] = current_dmr_data["num_ind"]
                ws[current_dmr_coll_fuel + current_dmr_row] = current_dmr_data[
                    "num_fuel"
                ]
                ws[current_dmr_coll_deadwood + current_dmr_row] = current_dmr_data[
                    "num_deadwood"
                ]

            del self.data_mapping["alp"][0:3]

    def write_att_data(self, wb):
        """Записываю книгу Титульный лист"""
        ws = wb["Титульный лист"]
        ws["O9"] = self.att_data["gplho"]
        ws["D9"] = self.att_data["enterprise"]
        ws["D10"] = self.att_data["forestry"]
        ws["D11"] = self.att_data["compartment"]
        ws["D12"] = self.att_data["sub_compartment"]
        ws["O11"] = float(self.att_data["area_square"])

    def run(self):
        file = BasicDir().get_basic_dir("templates/template_restatement.xlsx")
        workbook = load_workbook(file)

        self.write_trees_data(wb=workbook)
        self.write_att_data(wb=workbook)

        workbook.save(self.export_file)
        self.signal_status.emit(
            {
                "head": "Успешно",
                "body": "Данные успешно экспортированы в\n" + self.export_file,
            }
        )
