"""TODO: Исправить директорию сохранения, проверка открыт ли файл"""
import openpyxl
import os
from openpyxl.styles import Alignment, Font
from openpyxl.styles.borders import Border, Side
from openpyxl.worksheet import page
from PyQt5 import QtCore
import traceback

from ....src.models.nri import *


class Export_restatement_xls(QtCore.QThread):
    signal_status = QtCore.pyqtSignal(list)

    def __init__(self, **args):
        QtCore.QThread.__init__(self)
        self.all_data = args['data']
        self.dest_filename = args['path']

        self.wb = openpyxl.Workbook()
        self.ws0 = self.wb.create_sheet('ws0', 0)
        self.ws0.title = "Ведомость перечета деревьев"
        self.ws0.page_setup = page.PrintPageSetup(orientation='landscape', paperSize='9')  #Настройки печати(ориентация)
        self.ws1 = self.wb.create_sheet('ws1', 1)
        self.ws1.title = "Модельные деревья"
        self.ws1.page_setup = page.PrintPageSetup(orientation='landscape', paperSize='9')  #Настройки печати(ориентация)
        self.ws2 = self.wb.create_sheet('ws2', 2)
        self.ws2.title = "Количество деревьев по породам"
        self.ws2.page_setup = page.PrintPageSetup(orientation='landscape', paperSize='9')  #Настройки печати(ориентация)
        self.wb.remove_sheet(self.wb.get_sheet_by_name('Sheet'))

    def run(self):
        try:
            restatement = WSRestatement(worksheet=self.ws2, data=self.all_data).make()
            statement = WSStatement(worksheet=self.ws0, data=self.all_data).make()
            if restatement and statement:
                """Сохранение и открытие документа"""
                self.wb.save(filename=self.dest_filename)
                try:
                    os.system('start excel.exe ' + self.dest_filename)
                except:
                    None
        except Exception as error:
            self.signal_status.emit(['Ошибка сохранения', error, str(traceback.format_exc())])


class WSStatement:
    """ Ведомость перечёта деревьев, назначеных в рубку. Первый лист документа"""
    def __init__(self, **args):
        self.ws = args['worksheet']
        self.all_data = args['data']

    def make(self):
        alf = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
               'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
        for i in alf:
            self.ws.column_dimensions[i].width = 6.2

        """Применяю стиль ко всем ячейкам"""
        for r in range(1, 25):
            for c in range(1, 22):
                self.ws.cell(row=r, column=c).font = Font(size=14, name='Times New Roman')

        self.ws.cell(column=1, row=3, value='ВЕДОМОСТЬ ПЕРЕЧЕТА ДЕРЕВЬЕВ, НАЗНАЧЕНЫХ В РУБКУ')
        self.ws.merge_cells(start_row=3, end_row=3, start_column=1, end_column=21)
        self.ws.cell(row=3, column=1).alignment = Alignment(horizontal='center', vertical='center')
        self.ws.cell(row=3, column=1).font = Font(size=14, name='Times New Roman', bold=True)

        self.ws.merge_cells(start_row=4, end_row=5, start_column=1, end_column=21)

        self.ws.cell(column=1, row=6, value='Лесхоз:')
        self.ws.merge_cells(start_row=6, end_row=6, start_column=1, end_column=2)
        self.ws.cell(column=3, row=6, value=self.all_data['system']['enterprise'])  # Название лесхоза
        self.ws.cell(column=3, row=6).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=6, end_row=6, start_column=3, end_column=9)

        self.ws.cell(column=11, row=6, value='Лесничество:')
        self.ws.merge_cells(start_row=6, end_row=6, start_column=11, end_column=13)
        self.ws.cell(column=14, row=6, value=self.all_data['system']['forestry'])  # Название лесничества
        self.ws.cell(column=14, row=6).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=6, end_row=6, start_column=14, end_column=21)

        self.ws.cell(column=1, row=7, value='Вид пользования:')
        self.ws.merge_cells(start_row=7, end_row=7, start_column=1, end_column=4)
        self.ws.cell(column=5, row=7, value=Kind_use.get(Kind_use.code_kind_use == self.all_data['ta_fund']['code_kind_use']).name_kind_use)  # Вид пользования
        self.ws.cell(column=5, row=7).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=7, end_row=7, start_column=5, end_column=12)

        self.ws.cell(column=14, row=7, value='Вид рубки:')
        self.ws.merge_cells(start_row=7, end_row=7, start_column=14, end_column=16)
        self.ws.cell(column=17, row=7, value=Wct_meth.get(Wct_meth.code_wct_meth==self.all_data['ta_fund']['code_wct_meth']).name_wct_meth)  # Вид рубки
        self.ws.cell(column=17, row=7).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=7, end_row=7, start_column=17, end_column=21)

        self.ws.cell(column=1, row=8, value='Способ рубки:')
        self.ws.merge_cells(start_row=8, end_row=8, start_column=1, end_column=3)
        # self.ws.cell(column=4, row=8, value=!!!!)  # Способ рубки
        self.ws.cell(column=4, row=8).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=8, end_row=8, start_column=4, end_column=21)

        self.ws.cell(column=1, row=9, value='Квартал №')
        self.ws.merge_cells(start_row=9, end_row=9, start_column=1, end_column=3)
        self.ws.cell(column=4, row=9, value=self.all_data['system']['num_compartment'])  # Номер квартала
        self.ws.cell(column=4, row=9).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=9, end_row=9, start_column=4, end_column=5)
        self.ws.cell(column=4, row=9).alignment = Alignment(horizontal='left')

        self.ws.cell(column=6, row=9, value='Выдел №')
        self.ws.merge_cells(start_row=9, end_row=9, start_column=6, end_column=7)
        self.ws.cell(column=8, row=9, value=self.all_data['system']['num_sub_compartment'])  # Номер выдела
        self.ws.cell(column=8, row=9).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=9, end_row=9, start_column=8, end_column=10)
        self.ws.cell(column=8, row=9).alignment = Alignment(horizontal='left')

        self.ws.cell(column=11, row=9, value='делянка №')
        self.ws.merge_cells(start_row=9, end_row=9, start_column=11, end_column=13)
        # self.ws.cell(column=14, row=9, value=!!!!)  # Номер делянки
        self.ws.cell(column=14, row=9).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=9, end_row=9, start_column=14, end_column=16)
        self.ws.cell(column=14, row=9).alignment = Alignment(horizontal='left')

        self.ws.cell(column=17, row=9, value='Площадь')
        self.ws.merge_cells(start_row=9, end_row=9, start_column=17, end_column=18)
        self.ws.cell(column=19, row=9, value=self.all_data['system']['area'])  # площадь
        self.ws.cell(column=19, row=9).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=9, end_row=9, start_column=19, end_column=20)
        self.ws.cell(column=19, row=9).alignment = Alignment(horizontal='left')
        self.ws.cell(column=21, row=9, value='га.')

        self.ws.cell(column=1, row=10, value='в том числе не эксплуатационнная')
        self.ws.merge_cells(start_row=10, end_row=10, start_column=1, end_column=7)
        # self.ws.cell(column=8, row=10, value=!!!!)  # площадь не эксплуатационная
        self.ws.cell(column=8, row=10).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=10, end_row=10, start_column=8, end_column=15)
        self.ws.cell(column=8, row=10).alignment = Alignment(horizontal='left')
        self.ws.cell(column=16, row=10, value='га.')

        self.ws.cell(column=1, row=11, value='Лесосека №')
        self.ws.merge_cells(start_row=11, end_row=11, start_column=1, end_column=3)
        self.ws.cell(column=4, row=11, value=self.all_data['system']['num_cutting_area'])  # лесосека
        self.ws.cell(column=4, row=11).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=11, end_row=11, start_column=4, end_column=7)
        self.ws.cell(column=4, row=11).alignment = Alignment(horizontal='left')

        self.ws.cell(column=8, row=11, value='года, группа лесов')
        self.ws.merge_cells(start_row=11, end_row=11, start_column=8, end_column=11)
        self.ws.cell(column=12, row=11, value=Gr_forest.get(Gr_forest.code_gr_forest == self.all_data['ta_fund']['code_gr_forest']).name_gr_forest)  # группа лесов
        self.ws.cell(column=12, row=11).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=11, end_row=11, start_column=12, end_column=14)
        self.ws.cell(column=12, row=11).alignment = Alignment(horizontal='left')

        self.ws.cell(column=15, row=11, value=', разряд такс:')
        self.ws.merge_cells(start_row=11, end_row=11, start_column=15, end_column=17)
        self.ws.cell(column=18, row=11, value=Rank_trf.get(Rank_trf.code_rank_trf == self.all_data['ta_fund']['code_rank_trf']).name_rank_trf)  # разряд такс
        self.ws.cell(column=18, row=11).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=11, end_row=11, start_column=18, end_column=21)
        self.ws.cell(column=18, row=11).alignment = Alignment(horizontal='left')

        self.ws.cell(column=1, row=12, value='Категория защитности:')
        self.ws.merge_cells(start_row=12, end_row=12, start_column=1, end_column=5)
        # self.ws.cell(column=6, row=12, value=!!!!)  # Категория защитности
        self.ws.cell(column=6, row=12).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=12, end_row=12, start_column=6, end_column=21)
        self.ws.cell(column=6, row=12).alignment = Alignment(horizontal='left')

        self.ws.cell(column=1, row=13, value='Тип леса:')
        self.ws.merge_cells(start_row=13, end_row=13, start_column=1, end_column=2)
        self.ws.cell(column=3, row=13, value=Type_for.get(Type_for.code_type_for==self.all_data['ta_fund']['code_type_for']).name_type_for)  # тип леса
        self.ws.cell(column=3, row=13).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=13, end_row=13, start_column=3, end_column=5)
        self.ws.cell(column=3, row=13).alignment = Alignment(horizontal='left')

        self.ws.cell(column=6, row=13, value=', состав насаждения')
        self.ws.merge_cells(start_row=13, end_row=13, start_column=6, end_column=9)
        # self.ws.cell(column=10, row=13, value=!!!!)  # Формула состава
        self.ws.cell(column=10, row=13).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=13, end_row=13, start_column=10, end_column=14)
        self.ws.cell(column=10, row=13).alignment = Alignment(horizontal='left')

        self.ws.cell(column=15, row=13, value=', преобладающая порода')
        self.ws.merge_cells(start_row=13, end_row=13, start_column=15, end_column=19)
        # self.ws.cell(column=20, row=13, value=!!!!)  # преоблодающая порода
        self.ws.cell(column=20, row=13).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=13, end_row=13, start_column=20, end_column=21)
        self.ws.cell(column=20, row=13).alignment = Alignment(horizontal='left')

        self.ws.cell(column=1, row=14, value='Полнота')
        self.ws.merge_cells(start_row=14, end_row=14, start_column=1, end_column=2)
        self.ws.cell(column=3, row=14, value=self.all_data['ta_fund']['density'])  # полнота
        self.ws.cell(column=3, row=14).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=14, end_row=14, start_column=3, end_column=5)
        self.ws.cell(column=3, row=14).alignment = Alignment(horizontal='center')

        self.ws.cell(column=6, row=14, value=', возраст')
        self.ws.merge_cells(start_row=14, end_row=14, start_column=6, end_column=7)
        self.ws.cell(column=8, row=14, value=self.all_data['ta_fund']['age'])  # возраст
        self.ws.cell(column=8, row=14).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=14, end_row=14, start_column=8, end_column=11)
        self.ws.cell(column=8, row=14).alignment = Alignment(horizontal='center')

        self.ws.cell(column=12, row=14, value='Состояние насаждения')
        self.ws.merge_cells(start_row=14, end_row=14, start_column=12, end_column=16)
        self.ws.cell(column=17, row=14, value=Status.get(Status.code_status==self.all_data['ta_fund']['code_status']).name_status)  # Cостав насаждения
        self.ws.cell(column=17, row=14).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=14, end_row=14, start_column=17, end_column=21)
        self.ws.cell(column=17, row=14).alignment = Alignment(horizontal='left')

        self.ws.cell(column=1, row=15).border = Border(bottom=Side(style='thin'))
        self.ws.merge_cells(start_row=15, end_row=15, start_column=1, end_column=21)

        return True


class WSRestatement:
    """Колличество деревьев по породам. Третий лист документа"""
    def __init__(self, **args):
        self.ws = args['worksheet']
        # self.att = args['att']
        self.all_data = args['data']
        # self.amount = self.att['res_amount']
        self.dmrs = []
        self.species = {}

    def make(self):
        alf = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J',
               'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']
        for i in alf:
            self.ws.column_dimensions[i].width = 6.5
        """Формирую каркас таблицы"""
        column_number = 2
        for record in self.all_data['ta_fund_enum']['restatement_data']['quality_data']:
            """Собираю все диаметры и породы(для пород присваю номер столбца)"""
            if record['dmr'] not in self.dmrs:
                self.dmrs.append(record['dmr'])
            if record['species'] not in self.species:
                self.species[record['species']] = column_number
                column_number += 2
        self.dmrs.sort()

        dm = self.ws.cell(column=1, row=1, value='Ступень толщины, см.')
        self.ws.merge_cells(start_row=1, end_row=3, start_column=1, end_column=1)

        cell = self.ws.cell(column=2, row=1, value='Количество деревьев по породам')
        self.ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=len(self.species) * 2 + 1)
        cell.alignment = Alignment(wrapText=True)

        for spc in self.species:
            column = self.species[spc]
            """Название породы"""
            self.ws.cell(column=column, row=2, value=Species.get(Species.name_species_latin == spc).name_species)
            self.ws.merge_cells(start_row=2, end_row=2, start_column=column, end_column=column + 1)

            # Деловая\дровяная
            self.ws.cell(column=column, row=3, value='Дел')
            self.ws.cell(column=column + 1, row=3, value='Дров')

        null_row = 4
        for dr in self.dmrs:
            self.ws.cell(column=1, row=null_row, value=dr)
            null_row += 1

        self.ws.cell(column=1, row=len(self.dmrs) + 4, value='Итого')

        """Границы ячеек"""
        thin_border = Border(left=Side(style='thin'),
                             right=Side(style='thin'),
                             top=Side(style='thin'),
                             bottom=Side(style='thin'))
        for i in range(1, self.ws.max_row + 1):
            for ii in range(1, self.ws.max_column + 1):
                self.ws.cell(row=i, column=ii).border = thin_border
                self.ws.cell(row=i, column=ii).alignment = Alignment(horizontal='center',
                                                                vertical='center')  # Центрирую надпись

        dm.alignment = Alignment(wrapText=True)

        for record in self.all_data['ta_fund_enum']['restatement_data']['quality_data']:
            self.ws.cell(column=1, row=self.dmrs.index(record['dmr'])+4, value=record['dmr'])
            if record['num_ind'] > 0:
                self.ws.cell(column=self.species[record['species']], row=self.dmrs.index(record['dmr'])+4, value=record['num_ind'])
            if record['num_fuel'] > 0:
                self.ws.cell(column=self.species[record['species']]+1, row=self.dmrs.index(record['dmr'])+4, value=record['num_fuel'])
        """Сумма"""
        for i in range(len(self.all_data['ta_fund_enum']['restatement_data_amount'])):
            self.ws.cell(column=i+2, row=len(self.dmrs)+4, value=self.all_data['ta_fund_enum']['restatement_data_amount'][i])

        return True